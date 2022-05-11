#!/usr/bin/env python
# coding: utf-8

#補助スピンを使わない&予約可能な時間枠分のみスピンを確保
#必要なPythonモジュールをインポート
import numpy  as np
import pandas as pd
from dwave.system import EmbeddingComposite,DWaveSampler,LeapHybridSampler
from neal import SimulatedAnnealingSampler
from pyqubo import Array, Constraint, Placeholder
import itertools

#excel出力用モジュールをインポート
import openpyxl
from openpyxl import Workbook, load_workbook
from openpyxl.styles.fonts import Font
from openpyxl.styles import borders
from openpyxl.styles.alignment import Alignment
from openpyxl.styles.fills import PatternFill

#-------------------- 各患者のID、診察の種類、診察時間、検査の種類、検査時間のデータフレーム --------------------
columns = ['ID', 'Examination', 'Examination_time', 'Inspection', 'Inspection_time']
patient_df = pd.DataFrame([
    ['1', 'A', 15, 'a', 30],
    ['2', 'B', 30, 'b', 30],
    ['3', 'C', 15, 'c', 15],
    ['4', 'B', 30, 'a', 30],
    ['5', 'A', 15, 'b', 30],
    ['6', 'B', 15, 'c', 15],
    ['7', 'C', 30, 'a', 30],
    ['8', 'A', 15, 'a', 30],
    ['9', 'B', 15, 'c', 15],
    ['10', 'B', 15, 'c', 15],
    ['11', 'B', 15, 'a', 30],
    ['12', 'B', 30, 'a', 30],
    ['13', 'C', 15, 'a', 30],
    ['14', 'B', 30, 'a', 30],
    ['15', 'A', 15, 'b', 30],
    ['16', 'A', 15, 'c', 15],
    ['17', 'C', 30, 'a', 30],
    ['18', 'A', 15, 'a', 30],
    ['19', 'C', 15, 'b', 30],
    ['20', 'C', 15, 'c', 15]
], columns=columns)

#n_exam: 診察の種類数、n_insp: 検査の種類数
n_exam = len(patient_df['Examination'].unique())
n_insp = len(patient_df['Inspection'].unique())

#数字の番号と検査の種類を紐付け
exam_dic = {}; insp_dic = {}
for i in range(n_exam):
    exam_dic[i] = patient_df['Examination'].unique()[i]
for i in range(n_insp):
    insp_dic[i] = patient_df['Inspection'].unique()[i]

#-------------------- 診察・予約可能時間帯のデータフレーム --------------------
n_patients = 12                     #患者の数
n_days = 3                          #日数
n_timewins_day = 13                 #1日の時間枠の数
n_timewins = n_days*n_timewins_day  #時間枠の数

patient_df = patient_df.head(n_patients)
print(patient_df)

wb = load_workbook("shift_DB.xlsx")
ws = wb["Sheet1"]

#診察i(A,B,C)、検査i(a,b,c)を受けられる時間帯の辞書型リスト
task_time_list = [15, 30]

index = {}
exam_pos_dic = [[] for i in range(len(task_time_list))]
insp_pos_dic = [[] for i in range(len(task_time_list))]
for i in range(len(task_time_list)):
    exam_pos_dic[i] = {}
    insp_pos_dic[i] = {}

for i in range(n_exam):
    sample_list = [[] for i in range(len(task_time_list))]
    index[i] = exam_dic[i]
    for j in range(n_timewins):
        if ws.cell(row=i+4, column=j+2).value == '◯':
            sample_list[0].append(j+1)
            for k in range(len(task_time_list)-1):
                if ws.cell(row=i+4, column=j+3+k).value != '-':
                    sample_list[k+1].append(j+1)
    # if 13 in sample_list[1]:
    #     sample_list[1].remove(13)
    # if 26 in sample_list[1]:
    #     sample_list[1].remove(26)
    for k in range(n_days):
        if (k+1)*n_timewins_day in sample_list[1]:
            sample_list[1].remove((k+1)*n_timewins_day)

    for j in range(len(task_time_list)):
        exam_pos_dic[j][index[i]] = sample_list[j]    #診察時間がXX分の場合の予約可能なリスト

for i in range(n_exam, n_exam+n_insp):
    sample_list = [[] for i in range(len(task_time_list))]
    index[i] = insp_dic[i-n_exam]
    for j in range(n_timewins):
        if ws.cell(row=i+4, column=j+2).value == '◯':
            sample_list[0].append(j+1)
            for k in range(len(task_time_list)-1):
                if ws.cell(row=i+4, column=j+3+k).value != '-':
                    sample_list[k+1].append(j+1)
    if 13 in sample_list[1]:
        sample_list[1].remove(13)
    if 26 in sample_list[1]:
        sample_list[1].remove(26)
    for j in range(len(task_time_list)):
        insp_pos_dic[j][index[i]] = sample_list[j]    #検査時間がXX分の場合の予約可能なリスト

# print("各診察・検査の予約可能な時間帯の辞書型リスト")
# print(exam_pos_dic, exam_pos_dic2)
# print(insp_pos_dic, insp_pos_dic2)
# print("診察Aの予約可能な時間帯のリスト")
# print(exam_pos_dic['A'])
# print("検査aの予約可能な時間帯のリスト")
# print(insp_pos_dic['a'])

#患者i(1~20)が各々の診察及び検査を受けられる時間帯の辞書型リスト（→この分だけスピン定義）
patient_exam_dic = {}; patient_insp_dic = {}
for task_time in task_time_list:
    for i in range(n_patients):
        if patient_df['Examination_time'][i] == task_time:
            patient_exam_dic[i] = exam_pos_dic[int(task_time/15)-1][patient_df['Examination'][i]]
        if patient_df['Inspection_time'][i] == task_time:
            patient_insp_dic[i] = insp_pos_dic[int(task_time/15)-1][patient_df['Inspection'][i]]

# print("患者i(1~20)が各々の診察及び検査を受けられる時間帯の辞書型リスト")
# for i in range(n_patients):
#     print(patient_exam_dic[i], patient_insp_dic[i])

#============================== 以下アニーリング定式化 ==============================
#スピン定義: 患者iが診察①/検査/診察②をどの時間枠(j)に予約するか (患者iの診察①/検査/診察②の予約可能な時間枠数分だけスピンを確保)
x1 = []; x2 = []; y = []
for i in range(n_patients):
    x1.append(Array.create('x1['+str(i)+']', shape=(len(patient_exam_dic[i])), vartype='BINARY'))
    x2.append(Array.create('x2['+str(i)+']', shape=(len(patient_exam_dic[i])), vartype='BINARY'))
    y.append(Array.create('y['+str(i)+']', shape=(len(patient_insp_dic[i])), vartype='BINARY'))

#制約①：各患者は診察①、検査、診察②を1回ずつ受ける
#----------------------------------------------------------------------------------------------------------
r1_w = Placeholder('r1_w')
H1 = 0
for i in range(n_patients):
    H1 += (sum(x1[i][j] for j in range(len(patient_exam_dic[i]))) - 1)**2\
        + (sum(x2[i][j] for j in range(len(patient_exam_dic[i]))) - 1)**2\
        + (sum(y[i][j] for j in range(len(patient_insp_dic[i]))) - 1)**2
#----------------------------------------------------------------------------------------------------------

#制約②：診察①→検査→診察②の順番に受ける
#----------------------------------------------------------------------------------------------------------
r2_w = Placeholder('r2_w')
H2 = 0

for task_time in [15, 30]:
    for i in range(n_patients):
        for j in range(len(patient_exam_dic[i])):
            for k in range(len(patient_insp_dic[i])):
                if patient_df['Examination_time'][i] == task_time:
                    if patient_insp_dic[i][k] <= patient_exam_dic[i][j] + int(task_time/15)-1:
                        H2 += x1[i][j]*y[i][k]
                if patient_df['Inspection_time'][i] == task_time:
                    if patient_exam_dic[i][j] <= patient_insp_dic[i][k] + int(task_time/15)-1:
                        H2 += y[i][k]*x2[i][j]
#----------------------------------------------------------------------------------------------------------

#制約③：同じ診察/検査の1つの予約可能時間枠には1人のみが予約出来る
#----------------------------------------------------------------------------------------------------------
r3_w = Placeholder('r3_w')
H3 = 0
H3_ex = 0

for k in range(n_timewins):
    h3_exam = 0; h3_insp = 0
    for i in range(n_patients):
        if k+1 in patient_exam_dic[i]:
            h3_exam += x1[i][patient_exam_dic[i].index(k+1)] + x2[i][patient_exam_dic[i].index(k+1)]
        if k+1 in patient_insp_dic[i]:
            h3_insp += y[i][patient_insp_dic[i].index(k+1)]
    H3 += h3_exam*(h3_exam-1) + h3_insp*(h3_insp-1)

for task_time in [30]:
    for i in range(n_patients):
        for j in range(n_patients):
            if i != j and patient_df['Examination'][i] == patient_df['Examination'][j] and patient_df['Examination_time'][i] == task_time:
                for k in patient_exam_dic[i]:
                    if k+int(task_time/15)-1 in patient_exam_dic[j]:
                        H3_ex += x1[i][patient_exam_dic[i].index(k)]*x1[j][patient_exam_dic[j].index(k+int(task_time/15)-1)]\
                            + x1[i][patient_exam_dic[i].index(k)]*x2[j][patient_exam_dic[j].index(k+int(task_time/15)-1)]\
                            + x2[i][patient_exam_dic[i].index(k)]*x2[j][patient_exam_dic[j].index(k+int(task_time/15)-1)]
        for j in range(n_patients):
            if i != j and patient_df['Inspection'][i] == patient_df['Inspection'][j] and patient_df['Inspection_time'][i] == task_time:
                for k in patient_insp_dic[i]:
                    if k+int(task_time/15)-1 in patient_insp_dic[j]:
                        H3_ex += y[i][patient_insp_dic[i].index(k)]*y[j][patient_insp_dic[j].index(k+int(task_time/15)-1)]
#----------------------------------------------------------------------------------------------------------

#制約④：1日のうちに前診察→検査→後診察を一通り実施する(日を跨いではならない)
#----------------------------------------------------------------------------------------------------------
r4_w = Placeholder('r4_w')
H4 = 0

pairs = []
for m in range(n_days):
    for n in range(n_days):
        if m < n:
            pairs.append((m, n))

for (m,n) in pairs:
    for i in range(n_patients):
        for j in range(len(patient_exam_dic[i])):
            for k in range(len(patient_insp_dic[i])):
                if m * n_timewins_day < patient_exam_dic[i][j] and patient_exam_dic[i][j] <= (m+1) * n_timewins_day and n * n_timewins_day < patient_insp_dic[i][k] and patient_insp_dic[i][k] <= (n+1) * n_timewins_day:
                    H4 += x1[i][j]*y[i][k]
                elif m * n_timewins_day < patient_insp_dic[i][k] and patient_insp_dic[i][k] <= (m+1) * n_timewins_day and n * n_timewins_day < patient_exam_dic[i][j] and patient_exam_dic[i][j] <= (n+1) * n_timewins_day:
                    H4 += x1[i][j]*y[i][k]
            for k in range(len(patient_exam_dic[i])):
                if m * n_timewins_day < patient_exam_dic[i][j] and patient_exam_dic[i][j] <= (m+1) * n_timewins_day and n * n_timewins_day < patient_exam_dic[i][k] and patient_exam_dic[i][k] <= (n+1) * n_timewins_day:
                    H4 += x1[i][j]*x2[i][k]
                elif m * n_timewins_day < patient_exam_dic[i][k] and patient_exam_dic[i][k] <= (m+1) * n_timewins_day and n * n_timewins_day < patient_exam_dic[i][j] and patient_exam_dic[i][j] <= (n+1) * n_timewins_day:
                    H4 += x1[i][j]*x2[i][k]
#----------------------------------------------------------------------------------------------------------

#目的関数①：各患者の待ち時間を均一化
#----------------------------------------------------------------------------------------------------------
r5_w = Placeholder('r5_w')
H5 = 0
for i in range(n_patients-1):
    H5 += ((sum(patient_exam_dic[i+1][j] * x2[i+1][j]  for j in range(len(patient_exam_dic[i+1]))) - sum(patient_exam_dic[i+1][j] * x1[i+1][j]  for j in range(len(patient_exam_dic[i+1]))))
        -  (sum(patient_exam_dic[i][j] * x2[i][j]  for j in range(len(patient_exam_dic[i]))) - sum(patient_exam_dic[i][j] * x1[i][j]  for j in range(len(patient_exam_dic[i])))))**2
#----------------------------------------------------------------------------------------------------------

#目的関数②：各患者の待ち時間を短縮化
#----------------------------------------------------------------------------------------------------------
r6_w = Placeholder('r6_w')
H6 = 0
for i in range(n_patients):
    H6 += (sum(patient_exam_dic[i][j] * x2[i][j]  for j in range(len(patient_exam_dic[i]))) - sum(patient_exam_dic[i][j] * x1[i][j]  for j in range(len(patient_exam_dic[i]))))**2
#----------------------------------------------------------------------------------------------------------

#最終的なハミルトニアン(各項の重み付き和)
H = r1_w * H1 + r2_w * H2 + r3_w * H3 + r3_w * H3_ex + r4_w * H4 + r5_w * H5 + r6_w * H6

# max_one_list=[]
# max_one =''
# for k in range(n_exam):
#     for i in range(len(exam_pos_dic2[exam_dic[k]])):
#         for m in patient_df[(patient_df['Examination'] == exam_dic[k])].index.tolist():
#             if patient_df['Examination_time'][m] == 30:
#                 # print(exam_dic[k], 'x1[%d][%d]::' % (m, i)+ 'x2[%d][%d]::' % (m, i), end=" ")
#                 max_one = max_one + 'x1[%d][%d]::' % (m, i)+ 'x2[%d][%d]::' % (m, i)
#         # print("")
#         max_one = max_one + '@'
# max_one = max_one.split('@')[:-1]
# # print(max_one)

# for i in range(len(max_one)):
#     max_one[i]=max_one[i].split('::')[:-1]
#     print("max_one[%d] ="%(i), [1, max_one[i]])
#     # print(len(max_one[i]))
#     if len(max_one[i]) != 0:
#         max_one_list.append([1, max_one[i]])
# print("max_one_list =", max_one_list)

#============================== 以下アニーリング求解 ==============================
#ハミルトニアンからQUBOへのコンパイル
model = H.compile()
qubo, offset = model.to_qubo(feed_dict={'r1_w': 10, 'r2_w': 10, 'r3_w': 100, 'r4_w': 20, 'r5_w': 0.01, 'r6_w': 0.01})

#アニーリング実行
sampler = SimulatedAnnealingSampler()
responses = sampler.sample_qubo(qubo, num_reads=100, num_sweeps=10000)
# print(responses)

#-------------------- アニーリング結果整形・表示 --------------------
# print("#-------------------- アニーリング結果整形・表示 --------------------")
Rlist = []
for s, e, o in responses.data(['sample', 'energy', 'num_occurrences']):
    Rlist.append(s)
# print(Rlist)

# dict = {0:"前診察", 1:"検査　", 2:"後診察"}
results_task = {}
results_sequence = {}
for i in range(n_patients):
    task_list = []
    sequence_list = []
    # print("患者%dの診察&検査時間枠のレコメンド" % (i+1))
    # print("診察"+patient_df['Examination'][i]+": ", end=" ")
    for j in range(len(patient_exam_dic[i])):
        # print(Rlist[0]['x1[%d][%d]' % (i,j)], end=" ")
        if Rlist[0]['x1[%d][%d]' % (i,j)] == 1:
            sequence_list.append(patient_exam_dic[i][j])
            task_list.append(patient_df['Examination'][i])
    # print("")
    # print("検査"+patient_df['Inspection'][i]+": ", end=" ")
    for j in range(len(patient_insp_dic[i])):
        # print(Rlist[0]['y[%d][%d]' % (i,j)], end=" ")
        if Rlist[0]['y[%d][%d]' % (i,j)] == 1:
            sequence_list.append(patient_insp_dic[i][j])
            task_list.append(patient_df['Inspection'][i])
    # print("")
    # print("診察"+patient_df['Examination'][i]+": ", end=" ")
    for j in range(len(patient_exam_dic[i])):
        # print(Rlist[0]['x2[%d][%d]' % (i,j)], end=" ")
        if Rlist[0]['x2[%d][%d]' % (i,j)] == 1:
            sequence_list.append(patient_exam_dic[i][j])
            task_list.append(patient_df['Examination'][i])
    # print("")
    results_task[i] = task_list
    results_sequence[i] = sequence_list
    # print(task_list)
    # print(sequence_list)
    # print("待ち時間 =", sequence_list[2]-sequence_list[0])

#-------------------- 制約充足チェック --------------------
print("#-------------------- 制約充足チェック --------------------")
print("各患者は診察、検査、診察を1回ずつ受ける:", end=" ")
frag = 0
for i in range(n_patients):
    frag += (sum(Rlist[0]['x1[%d][%d]' % (i,j)] for j in range(len(patient_exam_dic[i]))) - 1)**2\
          + (sum(Rlist[0]['x2[%d][%d]' % (i,j)] for j in range(len(patient_exam_dic[i]))) - 1)**2\
          + (sum(Rlist[0]['y[%d][%d]' % (i,j)] for j in range(len(patient_insp_dic[i]))) - 1)**2
print("succeeded") if frag == 0 else print("failed")

print("診察→検査→診察の順番に受ける:", end=" ")
frag = 0
for i in range(n_patients):
    for j in range(len(results_sequence[i])-1):
        if results_sequence[i][j] >= results_sequence[i][j+1]:
            frag += 1
print("succeeded") if frag == 0 else print("failed")

print("日跨り禁止:", end=" ")
frag = 0
for (m,n) in pairs:
    for i in range(n_patients):
        # print(m * n_timewins_day, n * n_timewins_day)
        if m * n_timewins_day < min(results_sequence[i]) and min(results_sequence[i]) <= (m+1) * n_timewins_day and n * n_timewins_day < max(results_sequence[i]) and max(results_sequence[i]) <= (n+1) * n_timewins_day:
        # if min(results_sequence[i]) <= n_timewins_day and n_timewins_day < max(results_sequence[i]):
            frag += 1
print("succeeded") if frag == 0 else print("failed")

print("1つの時間枠には1人のみが予約可能:", end=" ")
frag = 0
for k in range(len(index)):
    frag1 = []
    for i in range(n_patients):
        for j in range(len(results_sequence[i])):
            if results_task[i][j] == index[k]:
                frag1.append(results_sequence[i][j])
    if len(frag1) != len(set(frag1)):
        frag += 1
print("succeeded") if frag == 0 else print("failed")

#-------------------- Excel結果出力 --------------------
print("#-------------------- Excel結果出力 --------------------")
print(results_task)
print(results_sequence)

for row in ws:
    for cell in row:
        cell.fill = PatternFill(fill_type = None)

col = ['349c3c', 'fc840c', '1c7cb4', 'e4442c', 'a47ccc', 'c1ab05', '946c45', '008b8b', 'e95388', '7d7b83', 'a7d28d',
       'f7b977', 'bbdbf3', 'f19ca7', 'ece093', 'cab8d9', 'ea5532', 'e62f8b', '26499d', '9cbb1c']

# for k in range(len(index)):
#     # print("index =",index[k])
#     for i in range(n_patients):
#         # print("Examination_time =",patient_df['Examination_time'][i])
#         for j in range(3):
#             if results_task[i][j] == index[k]:
#                 ws.cell(row=k+4, column=results_sequence[i][j]+1).fill = PatternFill(patternType='solid', start_color=col[i], end_color=col[i])

for k in range(len(exam_dic)):
    for i in range(n_patients):
        for j in range(3):
            if results_task[i][j] == exam_dic[k]:
                ws.cell(row=k+4, column=results_sequence[i][j]+1).fill = PatternFill(patternType='solid', start_color=col[i], end_color=col[i])
                if patient_df['Examination_time'][i] == 30:
                    ws.cell(row=k+4, column=results_sequence[i][j]+2).fill = PatternFill(patternType='solid', start_color=col[i], end_color=col[i])
for k in range(len(insp_dic)):
    for i in range(n_patients):
        for j in range(3):
            if results_task[i][j] == insp_dic[k]:
                ws.cell(row=k+7, column=results_sequence[i][j]+1).fill = PatternFill(patternType='solid', start_color=col[i], end_color=col[i])
                if patient_df['Inspection_time'][i] == 30:
                    ws.cell(row=k+7, column=results_sequence[i][j]+2).fill = PatternFill(patternType='solid', start_color=col[i], end_color=col[i])

wb.save("shift_DB.xlsx")

#-------------------- 待ち時間 --------------------
print("#-------------------- 待ち時間 --------------------")

wait_length = [0 for x in range(len(results_sequence))]
wait_result = []
for i in range(len(results_sequence)):
    wait_length[i] = (max(results_sequence[i])-min(results_sequence[i]))*15

wait_result.append(max(wait_length))
wait_result.append(min(wait_length))
wait_result.append(sum(wait_length)/len(wait_length))

print(wait_length)
labels_index = [15, 30, 45, 60, 75, 90, 105, 120, 135, 150, 165]
labels_list = [0 for x in range(len(labels_index)+1)]

for i in wait_length:
        if i in labels_index:
            print(labels_index.index(i))
            labels_list[labels_index.index(i)]+=1
        else:
            labels_list[len(labels_index)]+=1
print(labels_list)

# print(wait_result)
print("待ち時間最大値: ", wait_result[0])
print("待ち時間最小値: ", wait_result[1])
print("待ち時間平均値: ", wait_result[2])